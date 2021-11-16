""" All of the codes for writing the solution to an excel file. """

import openpyxl


class default_excel:
    def open_WB(File_name):
        try:
            workbook = openpyxl.load_workbook(File_name)
        except:
            workbook = openpyxl.Workbook()
        return workbook
    
    
    def select_sheet(sheetname, workbook):
        return workbook[sheetname]
    
    
    def empty_sheet(sheet):
        """delete the values on the sheet"""
        for row in sheet: 
            for cell in row:
                cell.value = None 
    
    def select_empty_sheet(sheetname, workbook):
        """select a sheet and clear it if it is exists, otherwise create a new sheet  """
        if sheetname not in workbook.sheetnames:
            workbook.create_sheet(sheetname)
        else:
            default_excel.empty_sheet(workbook[sheetname])
        return workbook[sheetname]          

class subwrite:
    def write_sol_time(sheet, model, rw=1, cl=1):
        """Writes the optimum and the time to excel"""
        sheet.cell(row=rw, column=cl).value = "Obj"
        sheet.cell(row=rw, column=cl+1).value =  model.objective.value()
     
        sheet.cell(row=rw+1, column=cl).value = "Time"
        sheet.cell(row=rw+1, column=cl+1).value =  model.solutionTime 
        
    def write_vektor(sheet, vektor, rw=1, cl=1, name =""):
        k = 0
        for each in vektor:
            sheet.cell(row=rw+k, column=1).value = name + str(each)
            sheet.cell(row=rw+k, column=cl+1).value =  vektor[each]
            k += 1
            
     
class modWrite:
    def write_ONE(sheetname, File_name, model, Transition_rules, OP, Premiums,  nbr_types, max_nbr_of_claims, rws, cws):
        """" Opens and write solution to a file """
        
        workbook = default_excel.open_WB(File_name)
        sheet = default_excel.select_empty_sheet(sheetname, workbook)  
        
        #obj+time
        subwrite.write_sol_time(sheet, model, rws, cws)
        
        
        #OP
        subwrite.write_vektor(sheet, OP, rw=5, cl = 1,  name ="OP")
        
        #TR
        TR_start = 6+nbr_types
        subwrite.write_vektor(sheet, Transition_rules, rw=TR_start, cl = 1,  name ="TR")
        
        #premiums
        Pr_start = TR_start + max_nbr_of_claims +2
        subwrite.write_vektor(sheet, Premiums, rw=Pr_start, cl = 1, name = "PR")
                
        workbook.save(filename=File_name)
        
    
        
        

    def write_MRun(sheetname, File_name, type_model, 
                   Obj, Time, OP, Transition_rules, Premiums, nbr_types, max_nbr_of_claims, start = 1):
        workbook = default_excel.open_WB(File_name)
        if start == 1:
            sheet = default_excel.select_empty_sheet(sheetname, workbook)  
        else:
            sheet = default_excel.select_sheet(sheetname, workbook)  
        
        #class+obj+time
        MRunWrite.class_sol_time(sheet, Obj, Time,start)
        
        #OP
        MRunWrite.dict_write(sheet, OP, start, 5, "OP")
        #TR
        TR_start = 6+nbr_types
        MRunWrite.dict_list_write(sheet, Transition_rules, start,TR_start , "T")
        
        #Premiums
        Pr_start = TR_start + max_nbr_of_claims +2
        MRunWrite.dict_write(sheet, Premiums, start, Pr_start, "Pr")
        
        #Save
        workbook.save(filename=File_name)
        
class MRunWrite:
    def class_sol_time(sheet, Obj, Time, start):
        if start == 1:
            sheet.cell(row=1, column=1).value = "class"
            sheet.cell(row=2, column=1).value = "Obj"
            sheet.cell(row=3, column=1).value = "Time"
        
        c = 0
        for k in Obj:
            sheet.cell(row=1, column=1+start+c).value =  k
            sheet.cell(row=2, column=1+start+c).value = Obj[k]
            sheet.cell(row=3, column=1+start+c).value =  Time[k]
            c += 1
    
    def dict_write(sheet, dictionary, start, rw, namevar):
        Num_col = 0
        for  String_col in dictionary:
            Num_row = 0
            for String_row in dictionary[String_col]:
                sheet.cell(row=rw+Num_row, column=1).value =  namevar+"_"+str(String_row)
                sheet.cell(row=rw+Num_row, column=1+start+Num_col).value = dictionary[String_col][String_row]
                Num_row += 1
            Num_col += 1
        
    def dict_list_write(sheet, dictionary, start, rw, namevar):
        Num_col = 0
        for  String_col in dictionary:
            for Num_row in range(len(dictionary[String_col])):
                sheet.cell(row=rw+Num_row, column=1).value =  namevar+"_"+str(Num_row)
                sheet.cell(row=rw+Num_row, column=1+start+Num_col).value = dictionary[String_col][Num_row]
            Num_col += 1
            
        

     
        
        
            
            
            